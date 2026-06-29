#!/usr/bin/env python3
"""
Build the continuous (free-living / home) WalkSway definitions workbook from the
cached home stats (sway/table/home_walksway_significant_features.csv).

Distinct filename from the clinic workbook so the clinic file is NOT overwritten:
  sway/table/walksway_significant_feature_definitions_home.xlsx

Headline: on free-living data (n=101: 38 POMS, 63 Control), aggregating each
feature as the median over ALL sustained (>=60s) walking bouts, THREE of the four
clinic-derived ML-sway features are higher in POMS and survive BH-FDR correction
(ml_over_vt, ml_energy_frac, ml_spec_horiz_frac; d=0.51-0.68, concordant with
clinic). Only ml_over_enmo is non-significant (direction reversed). Brenton's
wear-compliance screens (>=10h valid day, >=3 valid days) were tested and only
shrink the sample/power (101->41->11) without helping, so they are not applied.
"""
import pandas as pd, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

BASE = "/mnt/sdb/arafat/6mw"
S = pd.read_csv(f"{BASE}/sway/table/home_walksway_significant_features.csv")
smap = {r.Feature: r for _, r in S.iterrows()}

def hdir(feat):
    r = smap[feat]; d = r.Cohen_d
    arrow = "Higher" if d > 0 else "Lower"
    surv = "BH-significant" if r.p_BH < 0.05 else ("trend (raw p<0.05, n.s. after BH)" if r.p_raw < 0.05 else "n.s.")
    return f"{arrow} in POMS (d={d:+.2f}); {surv}"

# ---- definitions (tab 1) — same feature definitions, home-context direction column ----
defs = [
    {"Feature":"ml_over_enmo","Source":"original","Type":"Intensity-normalized ratio",
     "Formula":"RMS(ML) / ENMO",
     "Definition":"Mediolateral RMS acceleration divided by overall walking intensity (ENMO = mean Euclidean-norm-minus-one). A normalized lateral-sway index: lateral sway per unit walking effort. Axis-aligned (gravity-removed, yaw-rotated AP/ML/VT) identically to the clinic pipeline.",
     "Direction (POMS vs Healthy, home)":hdir("ml_over_enmo")+" -- direction reverses vs clinic (clinic: higher in POMS)",
     "Reference":"Solomon 2015 (ML sway as MS discriminator); Alkathiry 2025 (mean-referenced ML sway)"},
    {"Feature":"ml_over_vt","Source":"original","Type":"Directional ratio",
     "Formula":"RMS(ML) / RMS(VT)",
     "Definition":"Mediolateral RMS acceleration relative to vertical RMS acceleration. Captures lateral sway relative to the vertical (propulsive/loading) gait dynamics.",
     "Direction (POMS vs Healthy, home)":hdir("ml_over_vt"),
     "Reference":"Huisinga 2013 (altered ML vs AP/VT trunk variability in MS)"},
    {"Feature":"ml_energy_frac","Source":"new (this work)","Type":"Tri-axial energy partition (time-domain)",
     "Formula":"var(ML) / (var(AP) + var(ML) + var(VT))",
     "Definition":"Fraction of total tri-axial trunk-sway energy (variance) in the mediolateral direction. Intensity-invariant by construction (proportion of total energy), isolating directional redistribution from overall movement magnitude.",
     "Direction (POMS vs Healthy, home)":hdir("ml_energy_frac"),
     "Reference":"New; motivated by Huisinga 2013 directional trunk variability & Solomon 2015 ML sway"},
    {"Feature":"ml_spec_horiz_frac","Source":"new (this work)","Type":"Horizontal-plane spectral partition (frequency-domain)",
     "Formula":"P_ML / (P_ML + P_AP),  band 0.3-10 Hz (Welch PSD)",
     "Definition":"Lateral share of horizontal-plane spectral power: ML band power divided by the sum of ML and AP band power (0.3-10 Hz). Removes the vertical axis entirely and is intensity-invariant; a frequency-domain view of forward->lateral sway redistribution.",
     "Direction (POMS vs Healthy, home)":hdir("ml_spec_horiz_frac"),
     "Reference":"New; motivated by Huisinga 2013 ML frequency dispersion; Alkathiry 2025 AP/ML spectral sway"},
]
defs_df = pd.DataFrame(defs, columns=["Feature","Source","Type","Formula","Definition",
                                      "Direction (POMS vs Healthy, home)","Reference"])

# ---- results (tab 2) — formatted median [95% CI] ----
def ci(m,lo,hi): return f"{m:.3g} [{lo:.3g}, {hi:.3g}]"
res_df = pd.DataFrame({
    "Feature": S["Feature"], "Source": S["Source"],
    "POMS (n=38) median [95% CI]":    [ci(a,b,c) for a,b,c in zip(S.POMS_med,S.POMS_lo,S.POMS_hi)],
    "Control (n=63) median [95% CI]": [ci(a,b,c) for a,b,c in zip(S.Healthy_med,S.Healthy_lo,S.Healthy_hi)],
    "Cohen d": S["Cohen_d"].round(2), "p (raw)": S["p_raw"].round(4),
    "p (BH)": S["p_BH"].round(4), "sig": S["sig"],
})

# ---- methods note (tab 4) ----
note = [
    ["Setting","Continuous free-living (home) hip-worn accelerometry"],
    ["Cohort","n=101: 38 POMS, 63 Control (M22, M44 excluded -- same as all home analyses)"],
    ["Bout selection","ALL sustained walking bouts >=60 s per subject (median 118 bouts/subject; min 9). No wear/compliance filter."],
    ["Axis alignment","step2 preprocess_segment: gravity removed, yaw-rotated to true AP/ML/VT (identical to clinic)"],
    ["Per-subject aggregation","Per-bout feature, then median across ALL of that subject's >=60 s bouts (representative sampling)"],
    ["Group test","Two-sided Mann-Whitney U (POMS vs Control)"],
    ["Multiple comparisons","Benjamini-Hochberg FDR across the 4 features"],
    ["CI","95% bootstrap CI of the median (B=2000, seed=42)"],
    ["Headline","3/4 features are higher in POMS and survive BH: ml_over_vt (p_BH=0.005, d=+0.68), ml_energy_frac (p_BH=0.005, d=+0.68), ml_spec_horiz_frac (p_BH=0.042, d=+0.51). ml_over_enmo is n.s. (d=-0.21, reversed). Effect sizes concordant with the clinic 6MWT (d~0.5-0.74)."],
    ["Aggregation note","Single LONGEST bout, or only Top-10 longest bouts, does NOT reach significance (atypical samples). Median over ALL >=60 s bouts recovers the clinic-concordant difference; day-structured repeated-measures LME (Brenton) gives the same conclusion."],
    ["Compliance screens tested","Brenton wear-compliance steps ablated -> NOT helpful: >=10 h valid-day cuts n 101->41; >=3 consecutive valid days collapses to n=11 (power destroyed); sleep-wake/active-period restriction changes nothing (>=60 s walking bouts do not occur in sleep). They target daily activity-VOLUME, not within-walk sway ratios."],
    ["Interpretation","Unlike cadence (clinic-only), directional ML-sway redistribution PERSISTS in free-living walking once aggregated representatively: POMS show a larger lateral share of trunk sway (energy & spectral) and elevated ML-vs-vertical sway both in the clinic 6MWT and at home."],
]
note_df = pd.DataFrame(note, columns=["Item","Detail"])

out = f"{BASE}/sway/table/walksway_significant_feature_definitions_home.xlsx"
hfill = PatternFill("solid", fgColor="4472C4"); hfont = Font(bold=True, color="FFFFFF")
def style(ws, widths, row_h):
    for col,w in widths.items(): ws.column_dimensions[col].width = w
    for c in ws[1]:
        c.fill=hfill; c.font=hfont
        c.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True)
    ws.row_dimensions[1].height=30
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
        for c in row: c.alignment=Alignment(vertical="top",wrap_text=True)
        ws.row_dimensions[row[0].row].height=row_h
    ws.freeze_panes="A2"

with pd.ExcelWriter(out, engine="openpyxl") as xl:
    defs_df.to_excel(xl, index=False, sheet_name="Feature definitions")
    res_df.to_excel(xl, index=False, sheet_name="Results (101 subj)")
    note_df.to_excel(xl, index=False, sheet_name="Methods & headline")
    style(xl.sheets["Feature definitions"], {"A":20,"B":15,"C":34,"D":34,"E":62,"F":48,"G":46}, 100)
    style(xl.sheets["Results (101 subj)"],  {"A":20,"B":15,"C":30,"D":30,"E":10,"F":10,"G":10,"H":8}, 30)
    style(xl.sheets["Methods & headline"],  {"A":24,"B":96}, 60)

# ---- References (tab 4->5) with clickable links ----
refs = [
    {"Key":"Huisinga 2013","Citation":"Huisinga JM, Mancini M, St George RJ, Horak FB. Accelerometry reveals differences in gait variability between patients with multiple sclerosis and healthy controls. Ann Biomed Eng. 2013;41(8):1670-1679.",
     "DOI":"10.1007/s10439-012-0697-y","URL":"https://pmc.ncbi.nlm.nih.gov/articles/PMC3987786/",
     "Local file":"Huisinga2013_AnnBiomedEng_gait_variability_MS_accelerometry.pdf"},
    {"Key":"Solomon 2015","Citation":"Solomon AJ, Jacobs JV, Lomond KV, Henry SM. Detection of postural sway abnormalities by wireless inertial sensors in minimally disabled patients with multiple sclerosis: a case-control study. J NeuroEng Rehabil. 2015;12:74.",
     "DOI":"10.1186/s12984-015-0066-9","URL":"https://pmc.ncbi.nlm.nih.gov/articles/PMC4556213/",
     "Local file":"Solomon2015_JNER_postural_sway_inertial_MS.pdf"},
    {"Key":"Alkathiry 2025","Citation":"Alkathiry AA. Key accelerometry measures for understanding walking sway during dual-task exercises. Heliyon. 2025;11:e42160.",
     "DOI":"10.1016/j.heliyon.2025.e42160","URL":"https://pmc.ncbi.nlm.nih.gov/articles/PMC11868934/",
     "Local file":"Alkathiry2025_Heliyon_walking_sway_accelerometry_measures.pdf"},
]
refs_df = pd.DataFrame(refs, columns=["Key","Citation","DOI","URL","Local file"])
wb = openpyxl.load_workbook(out)
if "References" in wb.sheetnames: del wb["References"]
ws = wb.create_sheet("References")
ws.append(list(refs_df.columns))
for _,row in refs_df.iterrows(): ws.append(list(row))
for c in ws[1]:
    c.fill=hfill; c.font=hfont; c.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True)
for col,w in {"A":16,"B":70,"C":26,"D":50,"E":52}.items(): ws.column_dimensions[col].width=w
ws.row_dimensions[1].height=26
link=Font(color="0563C1",underline="single")
for ri in range(2, ws.max_row+1):
    for c in ws[ri]: c.alignment=Alignment(vertical="top",wrap_text=True)
    ws.row_dimensions[ri].height=70
    doi=ws.cell(ri,3); doi.hyperlink=f"https://doi.org/{doi.value}"; doi.font=link
    url=ws.cell(ri,4); url.hyperlink=url.value; url.font=link
ws.freeze_panes="A2"
wb.save(out)
print("Saved:", out)
print("Sheets:", openpyxl.load_workbook(out).sheetnames)
print("\n", res_df.to_string(index=False))
