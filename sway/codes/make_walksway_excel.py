import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill

# ---- definitions (tab 1) ----
defs = [
    {"Feature":"ml_over_enmo","Source":"original","Type":"Intensity-normalized ratio",
     "Formula":"RMS(ML) / ENMO",
     "Definition":"Mediolateral RMS acceleration divided by overall walking intensity (ENMO = mean Euclidean-norm-minus-one). A normalized lateral-sway index: lateral sway per unit walking effort.",
     "Direction (POMS vs Healthy)":"Higher in POMS (d=+0.74) — more lateral sway per unit effort",
     "Reference":"Solomon 2015 (ML sway as MS discriminator); Alkathiry 2025 (mean-referenced ML sway)"},
    {"Feature":"ml_over_vt","Source":"original","Type":"Directional ratio",
     "Formula":"RMS(ML) / RMS(VT)",
     "Definition":"Mediolateral RMS acceleration relative to vertical RMS acceleration. Captures lateral sway relative to the vertical (propulsive/loading) gait dynamics.",
     "Direction (POMS vs Healthy)":"Higher in POMS (d=+0.53) — lateral sway elevated relative to vertical",
     "Reference":"Huisinga 2013 (altered ML vs AP/VT trunk variability in MS)"},
    {"Feature":"ml_energy_frac","Source":"new (this work)","Type":"Tri-axial energy partition (time-domain)",
     "Formula":"var(ML) / (var(AP) + var(ML) + var(VT))",
     "Definition":"Fraction of total tri-axial trunk-sway energy (variance) that is in the mediolateral direction. Intensity-invariant by construction (proportion of total energy), isolating directional redistribution from overall movement magnitude.",
     "Direction (POMS vs Healthy)":"Higher in POMS (d=+0.51) — larger share of sway energy is lateral",
     "Reference":"New; motivated by Huisinga 2013 directional trunk variability & Solomon 2015 ML sway"},
    {"Feature":"ml_spec_horiz_frac","Source":"new (this work)","Type":"Horizontal-plane spectral partition (frequency-domain)",
     "Formula":"P_ML / (P_ML + P_AP),  band 0.3-10 Hz (Welch PSD)",
     "Definition":"Lateral share of horizontal-plane spectral power: ML band power divided by the sum of ML and AP band power (0.3-10 Hz). Removes the vertical axis entirely and is intensity-invariant; a frequency-domain view of forward->lateral sway redistribution.",
     "Direction (POMS vs Healthy)":"Higher in POMS (d=+0.51) — lateral dominates horizontal-plane spectral power",
     "Reference":"New; motivated by Huisinga 2013 ML frequency dispersion; Alkathiry 2025 AP/ML spectral sway"},
]
defs_df = pd.DataFrame(defs, columns=["Feature","Source","Type","Formula","Definition",
                                      "Direction (POMS vs Healthy)","Reference"])

# ---- results (tab 2) — formatted median [95% CI] ----
r = pd.read_csv("/mnt/sdb/arafat/6mw/sway/table/walksway_significant_features.csv")
def ci(m,lo,hi): return f"{m:.3g} [{lo:.3g}, {hi:.3g}]"
res_df = pd.DataFrame({
    "Feature": r["Feature"],
    "Source": r["Source"],
    "POMS (n=50) median [95% CI]": [ci(a,b,c) for a,b,c in zip(r.POMS_med,r.POMS_lo,r.POMS_hi)],
    "Healthy (n=70) median [95% CI]": [ci(a,b,c) for a,b,c in zip(r.Healthy_med,r.Healthy_lo,r.Healthy_hi)],
    "Cohen d": r["Cohen_d"].round(2),
    "p (raw)": r["p_raw"].round(4),
    "p (BH)": r["p_BH"].round(4),
    "sig": r["sig"],
})

out = "/mnt/sdb/arafat/6mw/sway/table/walksway_significant_feature_definitions.xlsx"
hfill = PatternFill("solid", fgColor="4472C4"); hfont = Font(bold=True, color="FFFFFF")
def style(ws, widths, row_h):
    for col,w in widths.items(): ws.column_dimensions[col].width = w
    for c in ws[1]:
        c.fill=hfill; c.font=hfont
        c.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True)
    ws.row_dimensions[1].height=30
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
        for c in row: c.alignment=Alignment(vertical="top",wrap_text=True)
        ws.row_dimensions[row[0].row].height=row_h
    ws.freeze_panes="A2"

with pd.ExcelWriter(out, engine="openpyxl") as xl:
    defs_df.to_excel(xl, index=False, sheet_name="Feature definitions")
    res_df.to_excel(xl, index=False, sheet_name="Results (120 subj)")
    style(xl.sheets["Feature definitions"],
          {"A":20,"B":15,"C":34,"D":34,"E":62,"F":40,"G":46}, 92)
    style(xl.sheets["Results (120 subj)"],
          {"A":20,"B":15,"C":30,"D":30,"E":10,"F":10,"G":10,"H":8}, 30)
print("Saved 2 tabs:", out)
import openpyxl; wb=openpyxl.load_workbook(out); print("Sheets:", wb.sheetnames)
print("\n", res_df.to_string(index=False))

# ---- References (tab 3) with clickable links ----
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
refs = [
    {"Key":"Huisinga 2013","Citation":"Huisinga JM, Mancini M, St George RJ, Horak FB. Accelerometry reveals differences in gait variability between patients with multiple sclerosis and healthy controls. Ann Biomed Eng. 2013;41(8):1670-1679.",
     "DOI":"10.1007/s10439-012-0697-y","URL":"https://pmc.ncbi.nlm.nih.gov/articles/PMC3987786/",
     "Local file":"Huisinga2013_AnnBiomedEng_gait_variability_MS_accelerometry.pdf"},
    {"Key":"Solomon 2015","Citation":"Solomon AJ, Jacobs JV, Lomond KV, Henry SM. Detection of postural sway abnormalities by wireless inertial sensors in minimally disabled patients with multiple sclerosis: a case-control study. J NeuroEng Rehabil. 2015;12:74.",
     "DOI":"10.1186/s12984-015-0066-9","URL":"https://pmc.ncbi.nlm.nih.gov/articles/PMC4556213/",
     "Local file":"Solomon2015_JNER_postural_sway_inertial_MS.pdf"},
    {"Key":"Alkathiry 2025","Citation":"Alkathiry AA. Key accelerometry measures for understanding walking sway during dual-task exercises. Heliyon. 2025;11:e42160.",
     "DOI":"10.1016/j.heliyon.2025.e42160","URL":"https://pmc.ncbi.nlm.nih.gov/articles/PMC11868934/",
     "Local file":"Alkathiry2025_Heliyon_walking_sway_accelerometry_measures.pdf"},
]
refs_df = pd.DataFrame(refs, columns=["Key","Citation","DOI","URL","Local file"])
wb = openpyxl.load_workbook(out)
if "References" in wb.sheetnames: del wb["References"]
ws = wb.create_sheet("References")
ws.append(list(refs_df.columns))
for _,row in refs_df.iterrows(): ws.append(list(row))
# style header + widths
hfill=PatternFill("solid",fgColor="4472C4"); hfont=Font(bold=True,color="FFFFFF")
for c in ws[1]:
    c.fill=hfill; c.font=hfont; c.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True)
for col,w in {"A":16,"B":70,"C":26,"D":50,"E":52}.items(): ws.column_dimensions[col].width=w
ws.row_dimensions[1].height=26
link=Font(color="0563C1",underline="single")
for ri in range(2, ws.max_row+1):
    for c in ws[ri]: c.alignment=Alignment(vertical="top",wrap_text=True)
    ws.row_dimensions[ri].height=70
    doi=ws.cell(ri,3); doi.hyperlink=f"https://doi.org/{doi.value}"; doi.font=link      # DOI clickable
    url=ws.cell(ri,4); url.hyperlink=url.value; url.font=link                            # URL clickable
ws.freeze_panes="A2"
wb.save(out)
print("Added References tab. Sheets:", openpyxl.load_workbook(out).sheetnames)
